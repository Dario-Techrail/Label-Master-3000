"""
Modulo unificato per tutta la logica di business dell'applicazione.
Contiene tutte le classi per l'elaborazione dei dati, generazione di documenti,
gestione componenti e serial number.
"""

import json
import os
from typing import List, Dict, Optional, Union
from datetime import datetime
from pathlib import Path

# Librerie per Excel/PDF/Word
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib.units import cm, mm
from reportlab.lib import colors


# ============================================================================
# GESTIONE COMPONENTI E DATABASE
# ============================================================================

class GestioneComponenti:
    """Classe per gestire i componenti con persistenza su file JSON."""

    def __init__(self, file_componenti: str = "DB/componenti_database.json"):
        """
        Inizializza il gestore componenti.

        Args:
            file_componenti (str): Nome del file JSON per salvare i componenti
        """
        self.file_componenti = file_componenti
        self.componenti = self._carica_componenti()

    def _carica_componenti(self) -> List[Dict]:
        """
        Carica i componenti dal file JSON.
        Aggiunge automaticamente il campo 'inizio_indicizzazione_prefisso' ai componenti
        che non lo possiedono (migrazione automatica).

        Returns:
            List[Dict]: Lista di componenti
        """
        if os.path.exists(self.file_componenti):
            try:
                with open(self.file_componenti, 'r', encoding='utf-8') as f:
                    componenti = json.load(f)
                
                # Migrazione automatica: aggiungi il campo mancante ai componenti vecchi
                migrazione_necessaria = False
                for comp in componenti:
                    if 'inizio_indicizzazione_prefisso' not in comp:
                        comp['inizio_indicizzazione_prefisso'] = None
                        migrazione_necessaria = True
                
                # Se necessaria la migrazione, salva subito il file aggiornato
                if migrazione_necessaria:
                    self._salva_componenti_lista(componenti)
                
                return componenti
            except Exception as e:
                print(f"Errore nel caricamento dei componenti: {e}")
                return []
        return []
    
    def _salva_componenti_lista(self, componenti: List[Dict]):
        """Salva una lista di componenti nel file JSON (usato per migrazione)."""
        try:
            with open(self.file_componenti, 'w', encoding='utf-8') as f:
                json.dump(componenti, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Errore nel salvataggio della migrazione componenti: {e}")

    def _salva_componenti(self):
        """Salva i componenti nel file JSON."""
        try:
            with open(self.file_componenti, 'w', encoding='utf-8') as f:
                json.dump(self.componenti, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Errore nel salvataggio dei componenti: {e}")

    def aggiungi_componente(self, nome: str, code_12nc: str,
                           sn_iniziale: Optional[int] = None,
                           prefisso_tipo_scheda: Optional[str] = None,
                           indicizzazione: bool = True,
                           inizio_indicizzazione_prefisso: Optional[Union[int, List[int]]] = None) -> bool:
        """
        Aggiunge un nuovo componente.

        Args:
            nome (str): Nome del componente
            code_12nc (str): Codice 12NC
            sn_iniziale (Optional[int]): SN iniziale (opzionale)
            prefisso_tipo_scheda (Optional[str]): Prefisso tipo scheda (opzionale)
            indicizzazione (bool): Flag per indicizzazione
            inizio_indicizzazione_prefisso (Optional[int|List[int]]): Numero iniziale per indicizzazione prefisso (opzionale). Può essere un intero o una lista separata.

        Returns:
            bool: True se aggiunto con successo, False se già esistente
        """
        if self.cerca_componente_per_nome(nome):
            return False

        componente = {
            'nome': nome,
            'code_12nc': code_12nc,
            'sn_iniziale': sn_iniziale,
            'prefisso_tipo_scheda': prefisso_tipo_scheda,
            'indicizzazione': indicizzazione,
            'inizio_indicizzazione_prefisso': inizio_indicizzazione_prefisso
        }

        self.componenti.append(componente)
        self._salva_componenti()
        return True

    def modifica_componente(self, nome_originale: str, nuovo_nome: str,
                           code_12nc: str, sn_iniziale: Optional[int] = None,
                           prefisso_tipo_scheda: Optional[str] = None,
                           indicizzazione: bool = True,
                           inizio_indicizzazione_prefisso: Optional[Union[int, List[int]]] = None) -> bool:
        """
        Modifica un componente esistente.

        Args:
            nome_originale (str): Nome originale del componente
            nuovo_nome (str): Nuovo nome del componente
            code_12nc (str): Codice 12NC
            sn_iniziale (Optional[int]): SN iniziale
            prefisso_tipo_scheda (Optional[str]): Prefisso tipo scheda
            indicizzazione (bool): Flag per indicizzazione

        Returns:
            bool: True se modificato con successo, False altrimenti
        """
        for i, comp in enumerate(self.componenti):
            if comp['nome'] == nome_originale:
                self.componenti[i] = {
                    'nome': nuovo_nome,
                    'code_12nc': code_12nc,
                    'sn_iniziale': sn_iniziale,
                    'prefisso_tipo_scheda': prefisso_tipo_scheda,
                    'indicizzazione': indicizzazione,
                    'inizio_indicizzazione_prefisso': inizio_indicizzazione_prefisso
                }
                self._salva_componenti()
                return True
        return False

    def elimina_componente(self, nome: str) -> bool:
        """
        Elimina un componente.

        Args:
            nome (str): Nome del componente da eliminare

        Returns:
            bool: True se eliminato con successo, False altrimenti
        """
        lunghezza_iniziale = len(self.componenti)
        self.componenti = [c for c in self.componenti if c['nome'] != nome]

        if len(self.componenti) < lunghezza_iniziale:
            self._salva_componenti()
            return True
        return False

    def aggiorna_sn_iniziale(self, nome: str, nuovo_sn: int) -> bool:
        """
        Aggiorna solo il campo sn_iniziale di un componente esistente.

        Args:
            nome (str): Nome del componente
            nuovo_sn (int): Nuovo valore per sn_iniziale

        Returns:
            bool: True se aggiornato con successo, False altrimenti
        """
        for comp in self.componenti:
            if comp['nome'] == nome:
                comp['sn_iniziale'] = nuovo_sn
                self._salva_componenti()
                print(f"DEBUG GC: SN iniziale per {nome} aggiornato a {nuovo_sn} nel database")
                return True
        print(f"DEBUG GC: Componente {nome} non trovato per aggiornamento SN")
        return False

    def cerca_componente_per_nome(self, nome: str) -> Optional[Dict]:
        """
        Cerca un componente per nome.

        Args:
            nome (str): Nome del componente

        Returns:
            Optional[Dict]: Componente trovato o None
        """
        for comp in self.componenti:
            if comp['nome'] == nome:
                return comp
        return None

    def ottieni_tutti_componenti(self) -> List[Dict]:
        """
        Restituisce tutti i componenti.

        Returns:
            List[Dict]: Lista di tutti i componenti
        """
        return self.componenti.copy()

    def ottieni_nomi_componenti(self) -> List[str]:
        """
        Restituisce solo i nomi dei componenti.

        Returns:
            List[str]: Lista dei nomi
        """
        return [comp['nome'] for comp in self.componenti]


# ============================================================================
# GESTIONE SERIAL NUMBER
# ============================================================================

class GestoreSerialNumber:
    """
    Classe per la gestione dei Serial Number con persistenza dei dati.
    """

    def __init__(self, file_stato: str = "DB/serial_numbers_state.json"):
        """
        Inizializza il gestore dei serial number.

        Args:
            file_stato (str): Nome del file JSON per salvare lo stato
        """
        self.file_stato = file_stato
        self.stato = self._carica_stato()

    def _carica_stato(self) -> dict:
        """
        Carica lo stato salvato dal file JSON.

        Returns:
            dict: Stato con descrizioni e ultimi SN
        """
        if os.path.exists(self.file_stato):
            try:
                with open(self.file_stato, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"Errore nel caricamento dello stato: {e}")
                return {}
        return {}

    def _salva_stato(self):
        """Salva lo stato in file JSON."""
        try:
            with open(self.file_stato, 'w', encoding='utf-8') as f:
                json.dump(self.stato, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Errore nel salvataggio dello stato: {e}")

    def _get_mese_lettera(self, mese: int) -> str:
        """
        Converte il numero del mese in lettera.
        01:A, 02:B, ..., 12:L

        Args:
            mese (int): Numero del mese (1-12)

        Returns:
            str: Lettera corrispondente
        """
        lettere = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
        return lettere[mese - 1]

    def genera_serial_number(self, descrizione: str, sn_iniziale: int = None, code_12nc: str = None) -> str:
        """
        Genera un Serial Number (SN) con formato: MYY NNNNN
        M = Mese (A-L), YY = Anno (ultimi 2 cifre), NNNNN = Numero incrementale

        Args:
            descrizione (str): Descrizione del prodotto
            sn_iniziale (int): Serial number da cui iniziare (opzionale)
            code_12nc (str): CODE 12NC del prodotto (opzionale, per salvataggio)

        Returns:
            str: Serial Number generato nel formato "J25 00138"
        """
        ora = datetime.now()
        mese = ora.month
        anno = ora.year % 100

        mese_lettera = self._get_mese_lettera(mese)

        # Ottieni il numero incrementale
        # Se viene specificato un sn_iniziale, usalo sempre (priorità massima)
        if sn_iniziale is not None:
            numero_incrementale = sn_iniziale
            print(f"DEBUG GSN: Usando sn_iniziale specificato: {numero_incrementale}")
        elif descrizione in self.stato:
            ultimo_sn = self.stato[descrizione]['ultimo_sn']
            numero_incrementale = ultimo_sn + 1
            print(f"DEBUG GSN: Componente esiste, usando ultimo_sn + 1: {numero_incrementale}")
        else:
            numero_incrementale = 0
            print(f"DEBUG GSN: Componente nuovo, partendo da 0")

        # Genera il SN
        sn = f"{mese_lettera}{anno:02d} {numero_incrementale:05d}"

        # Aggiorna lo stato
        self.stato[descrizione] = {
            'ultimo_sn': numero_incrementale,
            'data_ultimo_utilizzo': ora.isoformat(),
            'code_12nc': code_12nc
        }
        self._salva_stato()

        return sn

    def get_ultimo_sn(self, descrizione: str) -> int:
        """
        Ottiene l'ultimo serial number utilizzato per una descrizione.

        Args:
            descrizione (str): Descrizione del prodotto

        Returns:
            int: Numero dell'ultimo SN, -1 se non trovato
        """
        if descrizione in self.stato:
            return self.stato[descrizione]['ultimo_sn']
        return -1

    def lista_descrizioni(self) -> dict:
        """
        Restituisce tutte le descrizioni salvate.

        Returns:
            dict: Descrizioni con i loro ultimi SN
        """
        return self.stato


# ============================================================================
# GENERAZIONE DOCUMENTI EXCEL
# ============================================================================

class GeneratoreExcel:
    """
    Classe per la generazione di documenti Excel per la gestione documentazionale
    di prodotti, etichette di prodotti e gestionale magazzino.
    """

    def __init__(self, gestione_componenti: GestioneComponenti = None):
        """
        Inizializza il generatore Excel.

        Args:
            gestione_componenti (GestioneComponenti): Gestore componenti per sincronizzazione SN
        """
        self.workbook = None
        self.gestore_sn = GestoreSerialNumber()
        self.gestione_componenti = gestione_componenti

    def crea_documento_bus(self, bolla_produzione: str, bolla_vendita: str, numero_bus: int,
                          nome_file: str = "documento_bus.xlsx", bus_iniziale: int = 1,
                          fornitore: str = "TECHRAIL"):
        """
        Genera un documento Excel con tabella contenente Bolla Produzione, Bolla Vendita e Bus.

        Args:
            bolla_produzione (str): Numero della bolla di produzione
            bolla_vendita (str): Numero della bolla di vendita
            numero_bus (int): Quantità di bus da generare
            nome_file (str): Nome del file Excel da salvare
            bus_iniziale (int): Numero del primo bus
            fornitore (str): Nome del fornitore

        Returns:
            str: Percorso del file generato
        """
        if not bolla_produzione or not bolla_vendita:
            raise ValueError("Bolla Produzione e Bolla Vendita sono campi obbligatori")

        if numero_bus <= 0:
            raise ValueError("Il numero di bus deve essere maggiore di 0")

        self.workbook = Workbook()
        sheet = self.workbook.active
        sheet.title = "Documento Bus"

        headers = ["Fornitore", "Bolla Produzione", "Bolla Vendita", "Bus", "Descrizione",
                  "Serial Number (SN)", "SN Fornitore"]
        for col_idx, header_text in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header_text)
            self._formatta_header(cell)

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 30
        sheet.column_dimensions['F'].width = 20

        for row_idx in range(1, numero_bus + 1):
            bus_number = bus_iniziale + row_idx - 1

            cell_fornitore = sheet.cell(row=row_idx + 1, column=1, value=str(fornitore))
            self._formatta_cella(cell_fornitore)

            cell_produzione = sheet.cell(row=row_idx + 1, column=2, value=str(bolla_produzione))
            self._formatta_cella(cell_produzione)

            cell_vendita = sheet.cell(row=row_idx + 1, column=3, value=str(bolla_vendita))
            self._formatta_cella(cell_vendita)

            cell_bus = sheet.cell(row=row_idx + 1, column=4, value=f"Bus {bus_number:02d}")
            self._formatta_cella(cell_bus)

            cell_descrizione = sheet.cell(row=row_idx + 1, column=5, value="")
            self._formatta_cella(cell_descrizione)

            cell_sn = sheet.cell(row=row_idx + 1, column=6, value="")
            self._formatta_cella(cell_sn)

        self.workbook.save(nome_file)
        return nome_file

    def crea_documento_con_componenti(self, bolla_produzione: str, bolla_vendita: str,
                                     numero_bus: int, componenti: list,
                                     nome_file: str = "documento_componenti.xlsx",
                                     bus_iniziale: int = 1, fornitore: str = "TECHRAIL"):
        """
        Genera un documento Excel con Bus e Componenti associati.

        Args:
            bolla_produzione: Numero della bolla di produzione
            bolla_vendita: Numero della bolla di vendita
            numero_bus: Quantità di bus da generare
            componenti: Lista di componenti con quantità
            nome_file: Nome del file Excel
            bus_iniziale: Numero del primo bus
            fornitore: Nome del fornitore

        Returns:
            str: Percorso del file generato
        """
        if not bolla_produzione or not bolla_vendita:
            raise ValueError("Bolla Produzione e Bolla Vendita sono campi obbligatori")

        if numero_bus <= 0:
            raise ValueError("Il numero di bus deve essere maggiore di 0")

        if not componenti or len(componenti) == 0:
            raise ValueError("Almeno un componente è obbligatorio")

        self.workbook = Workbook()
        sheet = self.workbook.active
        sheet.title = "Componenti per Bus"

        headers = ["Fornitore", "Bolla Produzione", "Bolla Vendita", "Descrizione", "CODE 12NC", "SN", "Bus", "Tipo Scheda", "SN Fornitore"]
        for col_idx, header_text in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header_text)
            self._formatta_header(cell)

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 40
        sheet.column_dimensions['E'].width = 18
        sheet.column_dimensions['F'].width = 15
        sheet.column_dimensions['G'].width = 12
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 15

        row_counter = 2

        componenti_sn_calcolati = []
        for componente in componenti:
            nome_componente = componente.get('nome', '')
            sn_iniziale = componente.get('sn_iniziale', None)
            prefisso_tipo_scheda = componente.get('prefisso_tipo_scheda', None)
            code_12nc = componente.get('code_12nc', None)

            print(f"DEBUG BL: Ricevuto componente {nome_componente} con sn_iniziale={sn_iniziale}")

            componente_memorizzato = nome_componente in self.gestore_sn.stato

            if componente_memorizzato:
                code_12nc_salvato = self.gestore_sn.stato[nome_componente].get('code_12nc')
                code_12nc_finale = code_12nc if code_12nc else code_12nc_salvato
            else:
                if not code_12nc:
                    raise ValueError(f"CODE 12NC è obbligatorio al primo inserimento del componente: {nome_componente}")
                code_12nc_finale = code_12nc

            # Se l'utente ha specificato un sn_iniziale, usalo sempre
            # Altrimenti usa l'ultimo SN + 1 se il componente esiste, altrimenti 0
            if sn_iniziale is not None:
                prossimo_sn = sn_iniziale
                print(f"DEBUG BL: Usando SN specificato dall'utente: {prossimo_sn}")
            elif nome_componente in self.gestore_sn.stato:
                prossimo_sn = self.gestore_sn.stato[nome_componente]['ultimo_sn'] + 1
                print(f"DEBUG BL: Usando ultimo SN + 1 dal database: {prossimo_sn}")
            else:
                prossimo_sn = 0
                print(f"DEBUG BL: Componente nuovo, partendo da 0")

            componenti_sn_calcolati.append({
                'nome': nome_componente,
                'quantita': componente.get('quantita', 1),
                'prossimo_sn': prossimo_sn,
                'prefisso_tipo_scheda': prefisso_tipo_scheda,
                'code_12nc': code_12nc_finale,
                'indicizzazione': componente.get('indicizzazione', True),
                'inizio_indicizzazione_prefisso': componente.get('inizio_indicizzazione_prefisso')
            })

        # Flag per tracciare se è il primo utilizzo del componente in questa generazione
        primo_utilizzo = {comp_info['nome']: True for comp_info in componenti_sn_calcolati}
        # Contatore globale per tracciare gli elementi generati per componente
        elemento_globale = {comp_info['nome']: 0 for comp_info in componenti_sn_calcolati}

        for bus_idx in range(numero_bus):
            bus_number = bus_iniziale + bus_idx
            for comp_info in componenti_sn_calcolati:
                nome_componente = comp_info['nome']
                quantita = comp_info['quantita']
                prossimo_sn = comp_info['prossimo_sn']
                prefisso_tipo_scheda = comp_info['prefisso_tipo_scheda']
                code_12nc = comp_info['code_12nc']

                for elemento_idx in range(quantita):
                    # Passa prossimo_sn solo al primo utilizzo assoluto del componente
                    if primo_utilizzo[nome_componente]:
                        print(f"DEBUG BL: Generando primo SN per {nome_componente} con valore iniziale {prossimo_sn}")
                        part_number = self.gestore_sn.genera_serial_number(
                            nome_componente,
                            prossimo_sn,
                            code_12nc
                        )
                        primo_utilizzo[nome_componente] = False
                        print(f"DEBUG BL: Generato SN: {part_number}, nuovo ultimo_sn: {self.gestore_sn.stato[nome_componente]['ultimo_sn']}")
                    else:
                        part_number = self.gestore_sn.genera_serial_number(
                            nome_componente,
                            None,
                            None
                        )
                        print(f"DEBUG BL: Generato SN successivo: {part_number}, ultimo_sn: {self.gestore_sn.stato[nome_componente]['ultimo_sn']}")

                    tipo_scheda = ""
                    if prefisso_tipo_scheda:
                        if comp_info.get('indicizzazione', True):
                            # Usa inizio_indicizzazione_prefisso se impostato, altrimenti parte da 1
                            # L'incremento è per bus: ogni bus inizia da capo
                            inizio_indic = comp_info.get('inizio_indicizzazione_prefisso')
                            # Supporta int (start+offset) oppure lista di valori (sequenza per elemento)
                            if isinstance(inizio_indic, list) and len(inizio_indic) > 0:
                                numero_indic = inizio_indic[elemento_idx % len(inizio_indic)]
                                print(f"DEBUG BL: Tipo Scheda con inizio_indic=list {inizio_indic}, elemento_idx={elemento_idx}, numero_indic={numero_indic}")
                            elif isinstance(inizio_indic, int):
                                numero_indic = inizio_indic + elemento_idx
                                print(f"DEBUG BL: Tipo Scheda con inizio_indic={inizio_indic}, elemento_idx={elemento_idx}, numero_indic={numero_indic}")
                            else:
                                numero_indic = elemento_idx + 1
                            tipo_scheda = f"{prefisso_tipo_scheda}{numero_indic}"
                            print(f"DEBUG BL: Costruito tipo_scheda='{tipo_scheda}' da prefisso='{prefisso_tipo_scheda}' e numero={numero_indic}")
                        else:
                            tipo_scheda = f"{prefisso_tipo_scheda}"
                            print(f"DEBUG BL: Tipo Scheda senza indicizzazione='{tipo_scheda}'")
                    else:
                        print(f"DEBUG BL: prefisso_tipo_scheda è None/vuoto per componente {nome_componente}")

                    cell = sheet.cell(row=row_counter, column=1, value=str(fornitore))
                    self._formatta_cella(cell)

                    cell = sheet.cell(row=row_counter, column=2, value=str(bolla_produzione))
                    self._formatta_cella(cell)

                    cell = sheet.cell(row=row_counter, column=3, value=str(bolla_vendita))
                    self._formatta_cella(cell)

                    cell = sheet.cell(row=row_counter, column=4, value=nome_componente)
                    self._formatta_cella(cell)

                    cell = sheet.cell(row=row_counter, column=5, value=code_12nc)
                    self._formatta_cella(cell)

                    cell = sheet.cell(row=row_counter, column=6, value=part_number)
                    self._formatta_cella(cell)

                    cell = sheet.cell(row=row_counter, column=7, value=f"BUS {bus_number}")
                    self._formatta_cella(cell)

                    cell = sheet.cell(row=row_counter, column=8, value=tipo_scheda)
                    self._formatta_cella(cell)
                    print(f"DEBUG BL: Scritto tipo_scheda='{tipo_scheda}' nella riga {row_counter}")

                    cell = sheet.cell(row=row_counter, column=9, value="")
                    self._formatta_cella(cell)

                    row_counter += 1

        # Sincronizza il database componenti con gli ultimi SN utilizzati
        if self.gestione_componenti:
            print(f"DEBUG BL: Sincronizzazione database componenti...")
            for componente in componenti:
                nome_componente = componente.get('nome', '')
                if nome_componente in self.gestore_sn.stato:
                    ultimo_sn = self.gestore_sn.stato[nome_componente]['ultimo_sn']
                    print(f"DEBUG BL: Componente {nome_componente}, ultimo_sn utilizzato: {ultimo_sn}, nuovo sn_iniziale: {ultimo_sn + 1}")
                    # Aggiorna il componente nel database con il nuovo sn_iniziale
                    comp_db = self.gestione_componenti.cerca_componente_per_nome(nome_componente)
                    if comp_db:
                        print(f"DEBUG BL: Componente trovato nel DB, aggiornamento in corso...")
                        # Usa il valore ricevuto dall'interfaccia (componenti) per
                        # preservare l'eventuale inizio_indicizzazione inserito dall'utente
                        self.gestione_componenti.modifica_componente(
                            nome_originale=nome_componente,
                            nuovo_nome=comp_db['nome'],
                            code_12nc=comp_db['code_12nc'],
                            sn_iniziale=ultimo_sn + 1,
                            prefisso_tipo_scheda=comp_db.get('prefisso_tipo_scheda'),
                            indicizzazione=comp_db.get('indicizzazione', True),
                            inizio_indicizzazione_prefisso=componente.get('inizio_indicizzazione_prefisso')
                        )
                        print(f"DEBUG BL: Componente {nome_componente} aggiornato nel DB con sn_iniziale={ultimo_sn + 1}")
                    else:
                        print(f"DEBUG BL: ERRORE - Componente {nome_componente} NON trovato nel DB!")
                else:
                    print(f"DEBUG BL: Componente {nome_componente} NON trovato in gestore_sn.stato")

        self.workbook.save(nome_file)
        return nome_file

    def _formatta_header(self, cell):
        """Formatta una cella di header."""
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _formatta_cella(self, cell):
        """Formatta una cella di dati."""
        cell.font = Font(size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ============================================================================
# ELABORAZIONE DATI E GENERAZIONE FILE
# ============================================================================

class DataProcessor:
    """Classe per l'elaborazione dei dati Excel."""

    @staticmethod
    def generate_csv_reg(input_file: Path, output_file: Path, selected_descriptions: List[str],
                        extra_fields: dict = None) -> int:
        """
        Genera un file CSV Reg filtrando per descrizioni selezionate.

        Args:
            input_file: Percorso del file Excel di input
            output_file: Percorso del file Excel di output
            selected_descriptions: Lista delle descrizioni da includere
            extra_fields: Campi extra da aggiungere (da interfaccia)

        Returns:
            int: Numero di righe generate
        """
        if not Path(input_file).exists():
            raise FileNotFoundError(f"File di input non trovato: {input_file}")

        if extra_fields is None:
            extra_fields = {}

        wb_input = load_workbook(input_file)
        ws_input = wb_input.active

        input_headers = [cell.value for cell in ws_input[1]]
        input_col_map = {header: idx + 1 for idx, header in enumerate(input_headers)}

        required_cols = ["Fornitore", "Descrizione", "CODE 12NC", "SN", "Bus", "Tipo Scheda"]
        for col in required_cols:
            if col not in input_col_map:
                raise ValueError(f"Colonna '{col}' non trovata nel file di input")

        wb_output = Workbook()
        ws_output = wb_output.active
        ws_output.title = "EtichetteBOX"

        # Definisci tutti i campi di output nell'ordine richiesto
        output_headers = [
            "CODE 12NC", "DESCRIZIONE", "SN", "SN Fornitore", "Codice MAC", 
            "CLIENTE", "Bolla Vendita Techrail", "Bolla Produzione", 
            "Bus", "Modello Pullman", "Tipo Scheda", "PW Schede", 
            "PATH Certificato SSH", "PATH Certificato OVPN", "IP_VPN", 
            "Ordine Acquisto", "SN1", "SN3", "Ente_Trasporto"
        ]
        for col_idx, header_value in enumerate(output_headers, start=1):
            cell = ws_output.cell(row=1, column=col_idx, value=header_value)
            DataProcessor._formatta_header(cell)

        output_row = 2
        desc_col_idx = input_col_map["Descrizione"]

        # Mapping tra nomi colonne input e output (per colonne che vengono dal file input)
        input_to_output = {
            "CODE 12NC": "CODE 12NC",
            "Descrizione": "DESCRIZIONE",
            "SN": "SN",
            "SN Fornitore": "SN Fornitore",
            "Codice MAC": "Codice MAC",
            "Bus": "Bus",
            "Tipo Scheda": "Tipo Scheda",
            "SN1": "SN1",
            "SN3": "SN3"
        }

        # Mapping tra nomi campi extra_fields e output (campi da interfaccia)
        extra_fields_to_output = {
            "CLIENTE": "CLIENTE",
            "Bolla Vendita Techrail": "Bolla Vendita Techrail",
            "Bolla Produzione": "Bolla Produzione",
            "Modello Pullman": "Modello Pullman",
            "PW Schede": "PW Schede",
            "PATH Certificato SSH": "PATH Certificato SSH",
            "PATH Certificato OVPN": "PATH Certificato OVPN",
            "IP_VPN": "IP_VPN",
            "Ordine Acquisto": "Ordine Acquisto",
            "Ente_Trasporto": "Ente_Trasporto"
        }

        for input_row in range(2, ws_input.max_row + 1):
            descrizione = ws_input.cell(row=input_row, column=desc_col_idx).value

            if descrizione and descrizione in selected_descriptions:
                # Prepara i dati per questa riga
                row_data = {}

                # Leggi i dati dal file input
                for input_col_name, output_col_name in input_to_output.items():
                    if input_col_name in input_col_map:
                        value = ws_input.cell(row=input_row, column=input_col_map[input_col_name]).value
                        row_data[output_col_name] = str(value).upper() if value else ""
                    else:
                        row_data[output_col_name] = ""

                # Aggiungi i dati dall'interfaccia (extra_fields)
                for extra_field_name, output_col_name in extra_fields_to_output.items():
                    value = extra_fields.get(extra_field_name, "")
                    row_data[output_col_name] = str(value).upper() if value else ""

                # Popola SN1 e SN3 ricavandoli dal campo SN (split sul primo spazio)
                # Comportamento: SN1 = prima parte; SN3 = seconda parte (se presente).
                # Se SN vuoto o non splittabile, popola solo SN1 e lascia SN3 vuoto.
                sn_full = row_data.get('SN', '')
                if sn_full:
                    # split al primo spazio
                    parts = str(sn_full).split(' ', 1)
                    row_data['SN1'] = parts[0].upper() if parts[0] else ''
                    if len(parts) > 1 and parts[1].strip():
                        row_data['SN3'] = parts[1].upper()
                    else:
                        row_data['SN3'] = ''
                else:
                    row_data['SN1'] = ''
                    row_data['SN3'] = ''

                # Scrivi i dati nella riga di output
                for col_idx, header_name in enumerate(output_headers, start=1):
                    cell_value = row_data.get(header_name, "")
                    cell = ws_output.cell(row=output_row, column=col_idx, value=cell_value)
                    DataProcessor._formatta_cella(cell)

                output_row += 1

        rows_generated = output_row - 2
        if rows_generated == 0:
            raise ValueError("Nessuna riga trovata con le descrizioni selezionate")

        # Configura le larghezze delle colonne
        column_widths = {
            'A': 18,  # CODE 12NC
            'B': 40,  # DESCRIZIONE
            'C': 15,  # SN
            'D': 15,  # SN Fornitore
            'E': 15,  # Codice MAC
            'F': 15,  # CLIENTE
            'G': 20,  # Bolla Vendita Techrail
            'H': 20,  # Bolla Produzione
            'I': 12,  # Bus
            'J': 20,  # Modello Pullman
            'K': 15,  # Tipo Scheda
            'L': 15,  # PW Schede
            'M': 25,  # PATH Certificato SSH
            'N': 25,  # PATH Certificato OVPN
            'O': 15,  # IP_VPN
            'P': 20,  # Ordine Acquisto
            'Q': 15,  # SN1
            'R': 15,  # SN3
            'S': 20   # Ente_Trasporto
        }
        for col_letter, width in column_widths.items():
            ws_output.column_dimensions[col_letter].width = width

        wb_output.save(output_file)
        return rows_generated

    @staticmethod
    def generate_import_gestionale(input_file: Path, output_file: Path, selected_descriptions: List[str],
                                   extra_fields: dict = None) -> int:
        """
        Genera un file Import Gestionale filtrando per descrizioni selezionate.

        Args:
            input_file: Percorso del file Excel di input
            output_file: Percorso del file Excel di output
            selected_descriptions: Lista delle descrizioni da includere
            extra_fields: Campi extra da aggiungere (da interfaccia)

        Returns:
            int: Numero di righe generate
        """
        if not Path(input_file).exists():
            raise FileNotFoundError(f"File di input non trovato: {input_file}")

        if extra_fields is None:
            extra_fields = {}

        wb_input = load_workbook(input_file)
        ws_input = wb_input.active

        input_headers = [cell.value for cell in ws_input[1]]
        input_col_map = {header: idx + 1 for idx, header in enumerate(input_headers)}

        required_cols = ["Fornitore", "Descrizione", "CODE 12NC", "SN", "Bus", "Tipo Scheda"]
        for col in required_cols:
            if col not in input_col_map:
                raise ValueError(f"Colonna '{col}' non trovata nel file di input")

        wb_output = Workbook()
        ws_output = wb_output.active
        ws_output.title = "ImportGestionale"

        # Definisci tutti i campi di output nell'ordine richiesto per Import Gestionale
        output_headers = [
            "Fornitore", "CODE 12NC", "Sigla", "SN", "SN1", "SN2", "SN3", "quantità",
            "CLIENTE", "Ente_Trasporto", "PW Schede", "Unità techrail 3::Sigla",
            "Data ordine", "DESCRIZIONE", "Modello Pullman", "PATH Certificato SSH",
            "Codice MAC", "Codice fornitore", "Bolla Vendita Techrail", "Ordine",
            "Bus", "PATH Certificato OVPN", "SN Fornitore", "Bolla Produzione",
            "Unità techrail 3::Tipo etichetta", "Sito", "Tipo Scheda", "IP_VPN",
            "Unità techrail 3::CB Codice 12NC", "CB matricola", "Data Ricezione",
            "Ordine Acquisto", "NUC CB", "Nota", "SISTEMA", "flag"
        ]
        for col_idx, header_value in enumerate(output_headers, start=1):
            cell = ws_output.cell(row=1, column=col_idx, value=header_value)
            DataProcessor._formatta_header(cell)

        output_row = 2
        desc_col_idx = input_col_map["Descrizione"]

        # Mapping tra nomi colonne input e output (per colonne che vengono dal file input)
        input_to_output = {
            "Fornitore": "Fornitore",
            "CODE 12NC": "CODE 12NC",
            "Descrizione": "DESCRIZIONE",
            "SN": "SN",
            "SN Fornitore": "SN Fornitore",
            "Codice MAC": "Codice MAC",
            "Bus": "Bus",
            "Tipo Scheda": "Tipo Scheda",
            "Sigla": "Sigla",
            "quantità": "quantità",
            "Codice fornitore": "Codice fornitore",
            "Ordine": "Ordine",
            "Unità techrail 3::Tipo etichetta": "Unità techrail 3::Tipo etichetta",
            "Sito": "Sito",
            "Unità techrail 3::CB Codice 12NC": "Unità techrail 3::CB Codice 12NC",
            "CB matricola": "CB matricola",
            "Data Ricezione": "Data Ricezione",
            "NUC CB": "NUC CB",
            "Nota": "Nota",
            "SISTEMA": "SISTEMA",
            "flag": "flag",
            "SN1": "SN1",
            "SN2": "SN2",
            "SN3": "SN3"
        }

        # Mapping tra nomi campi extra_fields e output (campi da interfaccia)
        extra_fields_to_output = {
            "CLIENTE": "CLIENTE",
            "Bolla Vendita Techrail": "Bolla Vendita Techrail",
            "Bolla Produzione": "Bolla Produzione",
            "Modello Pullman": "Modello Pullman",
            "PW Schede": "PW Schede",
            "PATH Certificato SSH": "PATH Certificato SSH",
            "PATH Certificato OVPN": "PATH Certificato OVPN",
            "IP_VPN": "IP_VPN",
            "Ordine Acquisto": "Ordine Acquisto",
            "Ente_Trasporto": "Ente_Trasporto",
            "Unità techrail 3::Sigla": "Unità techrail 3::Sigla",
            "Data ordine": "Data ordine"
        }

        for input_row in range(2, ws_input.max_row + 1):
            descrizione = ws_input.cell(row=input_row, column=desc_col_idx).value

            if descrizione and descrizione in selected_descriptions:
                # Prepara i dati per questa riga
                row_data = {}

                # Leggi i dati dal file input
                for input_col_name, output_col_name in input_to_output.items():
                    if input_col_name in input_col_map:
                        value = ws_input.cell(row=input_row, column=input_col_map[input_col_name]).value
                        row_data[output_col_name] = str(value).upper() if value else ""
                    else:
                        row_data[output_col_name] = ""

                # Aggiungi i dati dall'interfaccia (extra_fields)
                for extra_field_name, output_col_name in extra_fields_to_output.items():
                    value = extra_fields.get(extra_field_name, "")
                    row_data[output_col_name] = str(value).upper() if value else ""

                # Popola SN1, SN2 e SN3 ricavandoli dal campo SN (split sul primo spazio)
                # Comportamento: SN1 = prima parte; SN2 = vuoto; SN3 = seconda parte (se presente).
                sn_full = row_data.get('SN', '')
                if sn_full:
                    # split al primo spazio
                    parts = str(sn_full).split(' ', 1)
                    row_data['SN1'] = parts[0].upper() if parts[0] else ''
                    row_data['SN2'] = ''  # SN2 è sempre vuoto per ora
                    if len(parts) > 1 and parts[1].strip():
                        row_data['SN3'] = parts[1].upper()
                    else:
                        row_data['SN3'] = ''
                else:
                    row_data['SN1'] = ''
                    row_data['SN2'] = ''
                    row_data['SN3'] = ''

                # Scrivi i dati nella riga di output
                for col_idx, header_name in enumerate(output_headers, start=1):
                    cell_value = row_data.get(header_name, "")
                    cell = ws_output.cell(row=output_row, column=col_idx, value=cell_value)
                    DataProcessor._formatta_cella(cell)

                output_row += 1

        rows_generated = output_row - 2
        if rows_generated == 0:
            raise ValueError("Nessuna riga trovata con le descrizioni selezionate")

        # Configura le larghezze delle colonne (tutte 15 per uniformità, puoi personalizzare)
        for col_idx in range(1, len(output_headers) + 1):
            ws_output.column_dimensions[ws_output.cell(1, col_idx).column_letter].width = 15

        wb_output.save(output_file)
        return rows_generated

    @staticmethod
    def generate_etichettebox_excel(input_file: Path, output_file: Path, selected_descriptions: List[str]) -> int:
        """
        Genera file EtichetteBOX filtrando per descrizioni selezionate.

        Args:
            input_file: File Excel di input
            output_file: File Excel di output
            selected_descriptions: Descrizioni da includere

        Returns:
            int: Numero di righe generate
        """
        return DataProcessor.generate_csv_reg(input_file, output_file, selected_descriptions)

    @staticmethod
    def extract_unique_descriptions(input_file: Path) -> List[str]:
        """
        Estrae le descrizioni univoche dal file Excel.

        Args:
            input_file: Percorso del file Excel

        Returns:
            List[str]: Lista di descrizioni univoche
        """
        if not Path(input_file).exists():
            raise FileNotFoundError(f"File non trovato: {input_file}")

        wb = load_workbook(input_file, read_only=True)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        try:
            desc_col_idx = headers.index("Descrizione") + 1
        except ValueError:
            raise ValueError("Colonna 'Descrizione' non trovata nel file")

        descriptions = set()
        for row in range(2, ws.max_row + 1):
            desc = ws.cell(row=row, column=desc_col_idx).value
            if desc and desc.strip():
                descriptions.add(desc.strip())

        wb.close()
        return sorted(list(descriptions))

    @staticmethod
    def _formatta_header(cell):
        """Formatta una cella di header."""
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    @staticmethod
    def _formatta_cella(cell):
        """Formatta una cella di dati."""
        cell.font = Font(size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ============================================================================
# GENERAZIONE Etichette Naz
# ============================================================================

class PDFLabelGenerator:
    """Classe per generare Etichette Naz con layout automatico, immagine e testo"""

    PAGE_WIDTH, PAGE_HEIGHT = A4
    COLS = 4
    ROWS = 21

    MARGIN_LEFT = 5 * mm
    MARGIN_RIGHT = 5 * mm
    MARGIN_TOP = 10 * mm
    MARGIN_BOTTOM = 10 * mm

    SPACING_X = 3 * mm
    SPACING_Y = 0

    FONT_SIZE = 10
    IMAGE_WIDTH = 20 * mm
    IMAGE_MARGIN = 2 * mm

    @staticmethod
    def generate_pdf_labels(input_file, output_file, image_path, filter_enabled=False, selected_tipo_scheda=None,
                          repetitions=1, start_column=1, start_row=1, font_size=5, image_width_mm=10):
        """
        Genera un PDF con etichette contenenti immagine e testo.

        Args:
            input_file: File Excel di input
            output_file: File PDF di output
            image_path: Path immagine logo
            filter_enabled: Se filtrare per CODE 12NC
            selected_tipo_scheda: CODE 12NC selezionati (nota: nome variabile mantenuto per compatibilità)
            repetitions: Numero ripetizioni
            start_column: Colonna iniziale
            start_row: Riga iniziale
            font_size: Dimensione font
            image_width_mm: Larghezza immagine in mm

        Returns:
            Numero di etichette generate
        """
        df_input = pd.read_excel(input_file)
        df_input = df_input[pd.notna(df_input['Tipo Scheda'])].copy()

        if filter_enabled:
            if not selected_tipo_scheda:
                raise ValueError("Nessun CODE 12NC selezionato")

            # Normalizza CODE 12NC nel dataframe (rimuovi spazi, converti a stringa)
            df_input['CODE 12NC'] = df_input['CODE 12NC'].astype(str).str.strip()

            # Normalizza i CODE 12NC selezionati (rimuovi spazi, converti a stringa)
            selected_normalized = [str(code).strip() for code in selected_tipo_scheda]

            # Debug: mostra i valori disponibili vs selezionati
            print(f"DEBUG PDFLabel: CODE 12NC disponibili nel file: {df_input['CODE 12NC'].unique().tolist()}")
            print(f"DEBUG PDFLabel: CODE 12NC selezionati (normalizzati): {selected_normalized}")

            df_filtered = df_input[df_input['CODE 12NC'].isin(selected_normalized)].copy()

            if df_filtered.empty:
                raise ValueError(f"Nessuna riga corrisponde ai CODE 12NC selezionati. "
                               f"Disponibili: {df_input['CODE 12NC'].unique().tolist()}, "
                               f"Selezionati: {selected_normalized}")
        else:
            df_filtered = df_input.copy()

        labels = []
        for _, row in df_filtered.iterrows():
            pn = f"PN. {row['CODE 12NC']}"
            sn = f"S.N. {row['SN']}"
            line1 = f"{pn}  {sn}"

            bus_value = str(row['Bus']).strip()
            if not bus_value.upper().startswith("BUS"):
                line2 = f"BUS {bus_value} – {row['Tipo Scheda']}"
            else:
                line2 = f"{bus_value} – {row['Tipo Scheda']}"

            labels.append((line1, line2))

        if repetitions > 1:
            labels = labels * repetitions

        empty_count = (start_row - 1) * PDFLabelGenerator.COLS + (start_column - 1)
        labels = [("", "")] * empty_count + labels

        usable_width = (PDFLabelGenerator.PAGE_WIDTH - PDFLabelGenerator.MARGIN_LEFT -
                       PDFLabelGenerator.MARGIN_RIGHT - (PDFLabelGenerator.COLS - 1) * PDFLabelGenerator.SPACING_X)
        usable_height = (PDFLabelGenerator.PAGE_HEIGHT - PDFLabelGenerator.MARGIN_TOP -
                        PDFLabelGenerator.MARGIN_BOTTOM - (PDFLabelGenerator.ROWS - 1) * PDFLabelGenerator.SPACING_Y)

        label_width = usable_width / PDFLabelGenerator.COLS
        label_height = usable_height / PDFLabelGenerator.ROWS

        image_width = image_width_mm * mm

        try:
            img = ImageReader(str(image_path))
            img_original_width, img_original_height = img.getSize()
            img_scale = image_width / img_original_width
            img_height = img_original_height * img_scale
        except Exception as e:
            print(f"Errore caricamento immagine: {e}")
            img = None
            img_height = image_width

        c = canvas.Canvas(str(output_file), pagesize=A4)
        labels_per_page = PDFLabelGenerator.COLS * PDFLabelGenerator.ROWS

        for i, (line1, line2) in enumerate(labels):
            if i > 0 and i % labels_per_page == 0:
                c.showPage()

            page_index = i % labels_per_page
            col = page_index % PDFLabelGenerator.COLS
            row = page_index // PDFLabelGenerator.COLS

            x = PDFLabelGenerator.MARGIN_LEFT + col * (label_width + PDFLabelGenerator.SPACING_X)
            y = (PDFLabelGenerator.PAGE_HEIGHT - PDFLabelGenerator.MARGIN_TOP -
                 (row + 1) * label_height - row * PDFLabelGenerator.SPACING_Y)

            if line1 or line2:
                if img:
                    img_x = x + 2 * mm
                    img_y = y + (label_height - img_height) / 2
                    c.drawImage(img, img_x, img_y,
                               width=image_width,
                               height=img_height,
                               preserveAspectRatio=True)

                text_x = x + image_width + PDFLabelGenerator.IMAGE_MARGIN + 2 * mm
                line_spacing = 2 * mm
                total_text_height = font_size * 2 + line_spacing
                text_start_y = y + (label_height + total_text_height) / 2 - font_size

                c.setFont("Helvetica", font_size)
                c.drawString(text_x, text_start_y, line1)

                c.setFont("Helvetica-Bold", font_size)
                c.drawString(text_x, text_start_y - font_size - line_spacing, line2)

        c.save()
        return len(labels)



# ============================================================================
# GENERAZIONE Etichette Interne (PDF CON FORMATO WORD-LIKE)
# ============================================================================

class WordLabelGenerator:
    """Classe per generare etichette in PDF con formato a righe alternate bianco/nero"""

    PAGE_TOP_MARGIN = 0.25
    PAGE_BOTTOM_MARGIN = 0

    PAGE_LEFT_MARGIN = 0.1
    PAGE_RIGHT_MARGIN = -0.1

    ROW_HEIGHT = 1.45
    COLUMNS_PER_ROW = 8

    FONT_SIZE = 7.5

    @staticmethod
    def generate_word_labels(input_file, output_file, filter_enabled=False, selected_tipo_scheda=None,
                            *args, add_counter=True, add_black_labels=True):
        """
        Genera un PDF con le etichette in formato tabella A4 Portrait.

        Args:
            input_file: File Excel di input
            output_file: File PDF di output
            filter_enabled: Se filtrare per tipo scheda
            selected_tipo_scheda: Tipi scheda selezionati
            *args: Parametri posizionali legacy
            add_counter: Se aggiungere il contatore Bus
            add_black_labels: Se aggiungere le etichette con sfondo nero

        Returns:
            Numero di etichette generate
        """
        repetitions = args[0] if len(args) > 0 else 1
        start_column = args[1] if len(args) > 1 else 1
        start_row = args[2] if len(args) > 2 else 1

        df_input = pd.read_excel(input_file)

        if filter_enabled:
            if not selected_tipo_scheda:
                raise ValueError("Nessun tipo scheda selezionato")
            df_filtered = df_input[df_input['Tipo Scheda'].isin(selected_tipo_scheda)].copy()
            if df_filtered.empty:
                raise ValueError("Nessuna riga corrisponde ai tipi scheda selezionati")
        else:
            df_filtered = df_input.copy()

        combinations = df_filtered[['Bus', 'Tipo Scheda']].drop_duplicates().sort_values(by=['Bus', 'Tipo Scheda'])

        labels = []
        for _, row in combinations.iterrows():
            bus = row['Bus']
            tipo = row['Tipo Scheda']
            bus_str = str(bus).replace("BUS", "").strip()
            if add_counter:
                labels.append(f"BUS {bus_str} – {tipo}")
            else:
                labels.append(tipo)

        if repetitions > 1:
            labels = labels * repetitions

        empty_count = (start_row - 1) * WordLabelGenerator.COLUMNS_PER_ROW + (start_column - 1)
        #LABEL DIVISION: 
        black_labels = labels
        if empty_count > 0:
            labels = [""] * empty_count + labels

        righe_dati = []
        for i in range(0, len(labels), WordLabelGenerator.COLUMNS_PER_ROW):
            riga = labels[i:i + WordLabelGenerator.COLUMNS_PER_ROW]
            righe_dati.append(riga)
        #black labels senza spazi
        righe_dati_black = []
        for i in range(0, len(black_labels), WordLabelGenerator.COLUMNS_PER_ROW):
            riga = black_labels[i:i + WordLabelGenerator.COLUMNS_PER_ROW]
            righe_dati_black.append(riga)

        page_width, page_height = A4
        top_margin = WordLabelGenerator.PAGE_TOP_MARGIN * cm
        bottom_margin = WordLabelGenerator.PAGE_BOTTOM_MARGIN * cm
        left_margin = WordLabelGenerator.PAGE_LEFT_MARGIN * cm
        right_margin = WordLabelGenerator.PAGE_RIGHT_MARGIN * cm

        label_height = WordLabelGenerator.ROW_HEIGHT * cm
        cols = WordLabelGenerator.COLUMNS_PER_ROW
        usable_width = page_width - left_margin - right_margin
        label_width = usable_width / cols

        c = canvas.Canvas(str(output_file), pagesize=A4)
        c.setFont("Helvetica", WordLabelGenerator.FONT_SIZE)

        total_rows_per_page = int((page_height - top_margin - bottom_margin) // label_height)
        page_row_counter = 0

        def _draw_rows(rows_list, bg_is_black=False):
            nonlocal page_row_counter
            for riga in rows_list:
                if page_row_counter >= total_rows_per_page:
                    c.showPage()
                    c.setFont("Helvetica", WordLabelGenerator.FONT_SIZE)
                    page_row_counter = 0

                y = page_height - top_margin - (page_row_counter + 1) * label_height

                for col_idx in range(cols):
                    x = left_margin + col_idx * label_width
                    if col_idx < len(riga) and riga[col_idx] != "":
                        text = str(riga[col_idx])
                    else:
                        text = ""

                    if bg_is_black:
                        c.setFillColor(colors.HexColor('#000000'))
                        text_color = colors.white
                    else:
                        c.setFillColor(colors.HexColor('#FFFFFF'))
                        text_color = colors.black

                    c.rect(x, y, label_width, label_height, stroke=0, fill=1)

                    if text:
                        c.setFillColor(text_color)
                        text_x = x + 2 * mm
                        text_y = y + label_height / 2 - 5
                        c.drawString(text_x, text_y, text)

                page_row_counter += 1

        _draw_rows(righe_dati, bg_is_black=False)

        if add_black_labels:
            _draw_rows(righe_dati_black, bg_is_black=True)

        c.save()
        return len(labels)


# ============================================================================
# Merge Sort documenti
# ============================================================================




class ExcelMerger:
    def __init__(self, file_list: List[str]):
        if not file_list:
            raise ValueError("La lista dei file Excel è vuota.")
        self.file_list = file_list
        self.dataframes = []
        self.columns = None

    def _load_and_validate_files(self):
        """
        Carica i file Excel e verifica che abbiano tutti le stesse colonne.
        """
        for file_path in self.file_list:
            df = pd.read_excel(file_path, dtype=str)  # tutto come testo

            if self.columns is None:
                self.columns = list(df.columns)
            else:
                if list(df.columns) != self.columns:
                    raise ValueError(
                        f"Formato non coerente nel file {file_path}.\n"
                        f"Attese: {self.columns}\n"
                        f"Trovate: {list(df.columns)}"
                    )

            self.dataframes.append(df)

    def merge_and_sort(
        self,
        sort_by: str,
        output_file: str,
        ascending: bool = True
    ):
        """
        Unisce i file, ordina per una colonna e salva il risultato.
        """
        self._load_and_validate_files()

        if sort_by not in self.columns:
            raise ValueError(f"La colonna '{sort_by}' non esiste.")

        merged_df = pd.concat(self.dataframes, ignore_index=True)

        # Forza di nuovo tutto a stringa (extra sicurezza)
        merged_df = merged_df.astype(str)

        # Sostituisci 'nan' con stringa vuota per avere celle vuote nell'Excel
        merged_df = merged_df.replace('nan', '', regex=False)

        merged_df = merged_df.sort_values(
            by=sort_by,
            ascending=ascending,
            kind="mergesort"  # ordinamento stabile
        )

        merged_df.to_excel(output_file, index=False)

        return merged_df
